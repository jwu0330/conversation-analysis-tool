"""依研究規則補齊對照組逐題 K/C/R 與知識點標籤。

K（KnowledgePointStatus）三態：沒有知識點 / Single / Multiple，知識點上限 3 個。
C（CorrectnessStatus）：只看單題；僅明顯事實錯誤為 Incorrect，灰區一律 Correct。
R（RepetitionStatus）：同一學生任一知識點連續出現第 3 題起為 Repeated；中斷重算。

知識點採教材的 15 點封閉詞彙。既有逐題語意標註作為人工審閱基礎；明示詞與
直接同義詞優先，若超過 3 點，保留題目明示且較具體的 3 點，不擴張隱含意義。
"""
from __future__ import annotations

import os
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "data" / "對話分析" / "原始檔"
CONTROL_SHEET_INDEX = 6
SOLO_COMBINED_SHEET_INDEX = 10
MAX_KP = 3

KP_ORDER = [
    "檔案傳輸", "FTP", "P2P", "BBS", "Telnet", "遠端登入", "電子郵件",
    "IMAP", "POP3", "SMTP", "即時通訊", "IM", "網路電話", "VoIP", "Skype",
]

# 只收直接名稱、常見全名、教材內明確代稱；不以用途反推未出現的概念。
ALIASES = {
    "FTP": ["ftp", "filezilla"],
    "P2P": ["p2p", "點對點", "bittorrent", "utorrent", "emule", "ares galaxy", "種子", "做種"],
    "BBS": ["bbs", "電子佈告欄", "電子布告欄", "電子佈藍", "佈告欄系統", "布告欄系統", "ptt"],
    "Telnet": ["telnet"],
    "遠端登入": ["遠端登入", "遠端連線"],
    "電子郵件": ["電子郵件", "郵件", "寄信", "收信", "信箱", "gmail", "outlook"],
    "IMAP": ["imap", "網際網路郵件存取協定"],
    "POP3": ["pop3", "郵局傳輸協定"],
    "SMTP": ["smtp", "簡單郵件傳輸協定"],
    "即時通訊": ["即時通訊", "即時訊息"],
    "IM": ["instant messenger", "im軟體", "im 軟體"],
    "網路電話": ["網路電話"],
    "VoIP": ["voip", "voice over internet protocol"],
    "Skype": ["skype"],
    "檔案傳輸": ["檔案傳輸", "傳檔案", "傳送檔案", "交換檔案", "上傳/下載", "上傳下載"],
}

ANAPHORA = ("那", "它", "他", "所以", "詳細一些", "具體來說", "是如何", "為什麼", "為甚麼")


def split_labels(value) -> list[str]:
    if pd.isna(value) or not str(value).strip():
        return []
    return [x.strip() for x in str(value).replace("｜", "|").split("|") if x.strip() in KP_ORDER]


def explicit_labels(question: str) -> list[str]:
    raw = str(question or "")
    # 貼入選擇題時只取「主題」欄，避免把錯誤選項中的干擾概念也標成知識點。
    if "主題：" in raw:
        raw = raw.split("主題：", 1)[1].split("知識點層", 1)[0]
    text = raw.lower().replace(" ", "")
    hits = []
    for kp in KP_ORDER:
        for alias in ALIASES[kp]:
            if alias.lower().replace(" ", "") in text:
                hits.append(kp)
                break
    return hits


def cap_labels(question: str, labels: list[str]) -> list[str]:
    """明示且具體者優先；同時提到協定時，泛稱（如電子郵件）可被具體點取代。"""
    labels = list(dict.fromkeys(k for k in labels if k in KP_ORDER))
    explicit = explicit_labels(question)
    merged = list(dict.fromkeys(explicit + labels))
    if len(merged) <= MAX_KP:
        return merged
    generic = {"檔案傳輸", "遠端登入", "電子郵件", "即時通訊", "網路電話"}
    ranked = sorted(
        merged,
        key=lambda k: (k not in explicit, k in generic, KP_ORDER.index(k)),
    )
    return ranked[:MAX_KP]


def k_status(labels: list[str]) -> str:
    return "沒有知識點" if not labels else ("Single" if len(labels) == 1 else "Multiple")


def correctness(question: str) -> str:
    """保守編碼：疑問、開放題、灰區均 Correct；清單僅放人工確認的明顯錯誤。"""
    # 本批對照組未發現單句本身達「明顯事實錯誤」門檻的敘述。
    obvious_incorrect: set[str] = set()
    return "Incorrect" if str(question).strip() in obvious_incorrect else "Correct"


def encode_frame(frame: pd.DataFrame) -> pd.DataFrame:
    d = frame.copy()
    cols = list(d.columns)
    uid, question = cols[5], cols[6]
    k_col, c_col, r_col, labels_col = cols[7], cols[8], cols[9], cols[10]

    final_by_index: dict[int, list[str]] = {}
    source_col = cols[1]
    for _, idxs in d.groupby(uid, dropna=False).groups.items():
        previous: list[str] = []
        for i in sorted(idxs, key=lambda j: float(d.at[j, source_col])):
            q = str(d.at[i, question]).strip()
            labels = explicit_labels(q)
            # 明確承接詞／代名詞且本題沒有直接名稱時，才承接上一題知識點。
            if not labels and q.lower().startswith(tuple(x.lower() for x in ANAPHORA)):
                labels = previous
            labels = cap_labels(q, labels)
            final_by_index[i] = labels
            previous = labels
    final_labels = [final_by_index[i] for i in d.index]

    d[labels_col] = [" | ".join(v) if v else None for v in final_labels]
    d[k_col] = [k_status(v) for v in final_labels]
    d[c_col] = [correctness(q) for q in d[question]]

    # 同一 CUserID 依來源列號排序；任一知識點連續第 3 次起為 Repeated。
    d[r_col] = "New"
    for _, idxs in d.groupby(uid, dropna=False).groups.items():
        ordered = sorted(idxs, key=lambda i: float(d.at[i, source_col]))
        streak: dict[str, int] = {}
        for i in ordered:
            current = set(final_labels[d.index.get_loc(i)])
            streak = {kp: streak.get(kp, 0) + 1 for kp in current if kp in streak or kp in current}
            # 上式對新知識點會從 1 開始；未出現在本題的點已自然移除（中斷歸零）。
            d.at[i, r_col] = "Repeated" if any(n >= 3 for n in streak.values()) else "New"
    return d


def write_sheet(path: Path, encoded: pd.DataFrame) -> None:
    wb = load_workbook(path)
    ws = wb.worksheets[CONTROL_SHEET_INDEX]
    headers = {cell.value: cell.column for cell in ws[3] if cell.value}
    for row_offset, (_, row) in enumerate(encoded.iterrows(), start=4):
        for name in ("KnowledgePointStatus", "CorrectnessStatus", "RepetitionStatus", "InvolvedKnowledgeLabels"):
            if name in headers:
                ws.cell(row=row_offset, column=headers[name]).value = None if pd.isna(row[name]) else row[name]
    wb.save(path)


def write_solo_combined(path: Path, encoded: pd.DataFrame) -> None:
    """把同一批對照組 KCR 同步到 SOLO 的整合逐題表，不改變列數或其他欄位。"""
    wb = load_workbook(path)
    ws = wb.worksheets[SOLO_COMBINED_SHEET_INDEX]
    headers = {cell.value: cell.column for cell in ws[3] if cell.value}
    # 同一人可能重複相同問句，以出現次序配對，避免覆蓋錯列。
    def norm(value) -> str:
        return "".join(str(value).split())

    queues: dict[tuple[str, str], list[tuple[str, str, str, str | None]]] = {}
    for _, row in encoded.iterrows():
        key = (str(row["CUserID"]), norm(row["UserQuestion"]))
        queues.setdefault(key, []).append((
            row["KnowledgePointStatus"], row["CorrectnessStatus"], row["RepetitionStatus"],
            None if pd.isna(row["InvolvedKnowledgeLabels"]) else row["InvolvedKnowledgeLabels"],
        ))
    used: dict[tuple[str, str], int] = {}
    for excel_row in range(4, ws.max_row + 1):
        if ws.cell(excel_row, headers["組別"]).value != "對照組":
            continue
        key = (
            str(ws.cell(excel_row, headers["CUserID"]).value),
            norm(ws.cell(excel_row, headers["UserQuestion"]).value),
        )
        pos = used.get(key, 0)
        if key not in queues or pos >= len(queues[key]):
            raise ValueError(f"SOLO 整合表找不到對照組 KCR 配對：row={excel_row}, key={key}")
        values = queues[key][pos]
        used[key] = pos + 1
        for name, value in zip(
            ("KnowledgePointStatus", "CorrectnessStatus", "RepetitionStatus", "InvolvedKnowledgeLabels"), values
        ):
            ws.cell(excel_row, headers[name]).value = value
    # SOLO_Score 是 SOLO_Level 的固定量化，不依樣本或模型估計；明確寫值以避免
    # Excel 公式快取在 openpyxl 儲存後消失。
    score_map = {"P": 0, "U": 1, "M": 2, "R": 3, "EA": 4}
    for excel_row in range(4, ws.max_row + 1):
        level = ws.cell(excel_row, headers["SOLO_Level"]).value
        if level in score_map:
            ws.cell(excel_row, headers["SOLO_Score"]).value = score_map[level]
    wb.save(path)


def validate(d: pd.DataFrame) -> None:
    k, c, r, labels = d.columns[7], d.columns[8], d.columns[9], d.columns[10]
    counts = d[labels].map(split_labels).map(len)
    expected = counts.map(lambda n: "沒有知識點" if n == 0 else ("Single" if n == 1 else "Multiple"))
    assert d[k].equals(expected)
    assert counts.max() <= MAX_KP
    assert d[[k, c, r]].notna().all().all()
    assert set(d[c]) <= {"Correct", "Incorrect"}
    assert set(d[r]) <= {"New", "Repeated"}


def main() -> int:
    targets = [p for p in RAW_DIR.glob("*.xlsx") if p.name.startswith(("Bloom_", "SOLO_"))]
    if not targets:
        raise SystemExit("找不到原始 Bloom/SOLO Excel。")
    reference = None
    for path in targets:
        frame = pd.read_excel(path, sheet_name=CONTROL_SHEET_INDEX, header=2)
        encoded = encode_frame(frame)
        validate(encoded)
        write_sheet(path, encoded)
        if path.name.startswith("SOLO_"):
            write_solo_combined(path, encoded)
        reference = encoded
        print(f"已更新：{path.name}（{len(encoded)} 題）")
    assert reference is not None
    print("K:", reference.iloc[:, 7].value_counts().to_dict())
    print("C:", reference.iloc[:, 8].value_counts().to_dict())
    print("R:", reference.iloc[:, 9].value_counts().to_dict())
    print("知識點數:", reference.iloc[:, 10].map(split_labels).map(len).value_counts().sort_index().to_dict())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
